import sys
import json 
import mysql.connector
import zlib
import hashlib
from docx import Document
from CustomProperty import CustomProperty
from DocxComposer import DocxComposer
from docx.shared import Inches
from docx.shared import Cm
from openpyxl import Workbook
#from openpyxl.styles import PatternFill

count = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
lastchapter = 0

def ReadLicense():
    with open("./License.key","rb") as f:
        c=zlib.decompress(f.read())
        License=json.loads(c)
        key=License['key']
        del License['key']
        if(key!=hashlib.sha1(json.dumps(License,separators=(',', ':')).encode("utf8")).hexdigest()):
            License=None
    return License



def generate(cursor, filename, documentid, componentid):
    query = 'Select * from Items where document= ' + documentid + ' and component=' + componentid + ' Order by ordering asc;'
    cursor.execute(query)
    result = cursor.fetchall()

    document = Document('./plugins/template/template_'+documentid)
    DC = DocxComposer(cursor, document,documentid, componentid)
    for item in result:
        if(item[3] == 0):  # Document Itemmy
            DC.insertDOC(item)
        #elif(item[3] == 1): #SSS Requirement Item
        #    insertReq(document, item)
        elif(item[3] == 11):  # SRS Requirement Item
            DC.insertSRSReq(item)
        elif(item[3] == 21):  # SDD Requirement Item
            DC.insertSDDReq(item)
        elif(item[3] == 31):  # STD Requirement Item
            DC.insertSTDReq(item)
        else:
            None
    
    document.save("./gen/"+filename)
    CustomProperty(cursor, './gen/'+filename, documentid, componentid)



def generate_xls(cursor, filename, documentid, componentid, filter):
    wb = Workbook()
    ws = wb.active

    if(filter == "Document Only"):
        query = 'Select * from Items where type=0 and document= ' + documentid + ' and component=' + componentid + ' Order by ordering asc;'
    elif(filter == "Requirement Only"):
        query = 'Select * from Items where type='+documentid+'1 and document= ' + documentid + ' and component=' + componentid + ' Order by ordering asc;'
    else:
        query = 'Select * from Items where document= ' + documentid + ' and component=' + componentid + ' Order by ordering asc;'

    cursor.execute(query)
    result = cursor.fetchall()

    ws.append(['uid', "isReq", 'id', 'requirement','src','uc','description'])

    for item in result:
        if(item[3] == 0):  # Document item
            ws.append([item[0], " ", levelCounter(item[5]), item[6], item[7], item[8], item[9]])
        else:  # requirement Item
            ws.append([item[0], "o", item[5], item[6], item[7], item[8], item[9]])
    wb.save('./gen/'+filename)
    return './gen/'+filename

    
def generate_rof(cursor, filename, documentid, componentid, confirmed):
    wb = Workbook()
    ws = wb.active
    query = 'Select * from Items where document= ' + documentid + ' and component=' + componentid + ' Order by ordering asc;'
    cursor.execute(query)
    result = cursor.fetchall()
    level=dict()
    lv=""
    for item in result:
        if(item[3] == 0):  # Document item
            lv=levelCounter(item[5])
        level[item[0]]=lv
        if(item[3]!=0):
            level[item[0]]=lv+" "+item[5]
        

    query = 'Select rm.reviews.reviewer, rm.reviews.asis, rm.reviews.tobe, rm.reviews.comment, rm.reviews.state, rm.items.uid from rm.reviews left join rm.items on rm.reviews.itemid=rm.items.uid where rm.reviews.itemid in (select uid from rm.items where rm.items.document= ' + documentid + ' and rm.items.component=' + componentid + ' ) Order by rm.items.ordering asc;'
    #print(query)
    cursor.execute(query)
    result = cursor.fetchall()

    ws.append(['i','where', 'AS-IS', 'TO-BE', 'owner','Accept/Reject','comment'])

    for c,item in enumerate(result):
        STATE="Accept" if item[4]=="1" else "Reject" if item[4]=="2" else ""
        ws.append([str(c+1),level[item[5]], item[1], item[2],item[0],STATE,item[3]])
    wb.save('./gen/'+filename)
    #wb.save('./gen/'+filename.replace(".rof",".xlsx"))
    return './gen/'+filename

def levelCounter(lv):
    global lastchapter
    global count
    pos = (len(lv) // 2) - 1
    count[pos] = count[pos]+1
    if (lastchapter > pos):
        for i in range(pos+1, len(count)):
            count[i] = 0
        if (count[0] == 0):
            count[0] = 1
    lastchapter = pos
    level = [str(x) for x in count[0:pos+1]]
    return ".".join(level)

if __name__ == '__main__':
    db=sys.argv[1]
    filename=sys.argv[2]
    documentid=sys.argv[3]
    componentid=sys.argv[4]
    filter=sys.argv[5]
    
    configs=ReadLicense()
    del configs['mysql']['multipleStatements']
    configs['mysql']['database']=db

    cnx = mysql.connector.connect(**configs['mysql'])
    cursor = cnx.cursor()
    if('docx' in filename):
        generate(cursor, filename, documentid, componentid)
    elif('.rof' in filename):
        generate_rof(cursor, filename, documentid, componentid,0)
    else:
        generate_xls(cursor, filename, documentid, componentid, filter)
    cnx.close()

