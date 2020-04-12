import sys
import json
import zlib
import hashlib
import mysql.connector
from openpyxl import load_workbook


def ReadLicense():
	with open("./License.key","rb") as f:
		c=zlib.decompress(f.read())
		License=json.loads(c)
		key=License['key']
		del License['key']
		if(key!=hashlib.sha1(json.dumps(License,separators=(',', ':')).encode("utf8")).hexdigest()):
			License=None
	return License



def import_template(cursor, filename, documentid, componentid):
	wb = load_workbook(filename)
	ws=wb['Sheet']
	items=[]
	for c,row in enumerate(ws):
		if(c>0):
			items.append([cell.value for cell in row])
	query = 'Select ordering from items where document= '+documentid + ' and component = '+componentid+' Order by ordering desc limit 1;'
	cursor.execute(query)
	try:
		order=cursor.fetchone()[0]+1
	except:
		order=0

	for item in items:
		query = 'Insert into items (document, component, type, ordering, identifier, title, src, uc, description) VALUES('+documentid+','+componentid+','+('0' if item[1]!='o' else str(int(documentid)*10+1))+','+str(order)+',"'+((len(item[2])+1)//2*'1.' if item[1]!='o' else item[2])+'","'+(str(item[3]) if item[3]!=None else "")+'","'+(str(item[4].replace("\"","'")) if item[4]!=None else "")+'","'+(str(item[5].replace("\"","'")) if item[5]!=None else "")+'","'+(str(item[6].replace("\"","'")) if item[6]!=None else "")+'");'
		cursor.execute(query)
		order+=1
	cnx.commit()
	

def import_interface(cursor, filename, documentid, componentid):
	wb = load_workbook(filename)
	ws=wb['Sheet']
	items=[]
	for c,row in enumerate(ws):
		if(c>0):
			items.append([cell.value for cell in row])
	query = 'Select * from items where document= '+documentid + ' and component = '+componentid+' Order by ordering desc limit 1;'
	result = cursor.fetchall()
	delid=[]
	order=0
	for item in result:
		if("-IDR-" in item[5]):
			delid.append(item[0])
		if("인터페이스" in item[6] and item[5]=="1.1.1."):
			order=item[4]+0.0001

	for item in items:
		query = 'Insert into items (document, component, type, ordering, identifier, title, src, uc, description, testtype) VALUES('+documentid+','+componentid+','+('0' if item[1]!='o' else str(int(documentid)*10+1))+','+str(order)+',"'+((len(item[2])+1)//2*'1.' if item[1]!='o' else item[2])+'","'+(str(item[3]) if item[3]!=None else "")+'","'+(str(item[4].replace("\"","'")) if item[4]!=None else "")+'","'+(str(item[5].replace("\"","'")) if item[5]!=None else "")+'","'+(str(item[6].replace("\"","'")) if item[6]!=None else "")+'",1);'
		cursor.execute(query)
		order+=0.0001
	cnx.commit()


if __name__ == '__main__':
	filename=sys.argv[1]
	documentid=sys.argv[2]
	componentid=sys.argv[3]

	configs=ReadLicense()
	del configs['mysql']['multipleStatements']
	cnx = mysql.connector.connect(**configs['mysql'])
	cursor = cnx.cursor()
	if(".TEMPLATE" in filename)
		import_template(cursor, filename,documentid,componentid)
	else:
		import_interface(cursor, filename,documentid,componentid)
	cnx.close()
